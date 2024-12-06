import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import base64


def extract(d, keys):
    return [d[key] for key in keys]


def autosize_columns(ws, fixed={}):
    # Try to autosize the columns (doesn't always work due to dynamic
    # content, font family and font size differences, etc.) There is no
    # easy way to do this when buiding excel files.
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        try:
            l = get_column_letter(col[0].column)
            if l in fixed:
                adjusted_width = fixed["l"]
            ws.column_dimensions[l].width = adjusted_width
        except:
            l = col[0].column
            if l in fixed:
                adjusted_width = fixed["l"]
            ws.column_dimensions[l].width = adjusted_width


def create_spreadsheet(
    headers,
    rows,
    fileorbuffer,
    styles={},
    merge=[],
    formats={},
    named_styles=[],
    freeze_panes="A2",
    dimensions=None,
    auto_size=True,
):
    wb = openpyxl.Workbook()
    ws = wb.active

    local_styles = {}

    style = NamedStyle(name="col_header")
    style.font = Font(bold=True)
    style.border = Border(bottom=Side(style="medium", color="000000"))
    local_styles[style.name] = style

    style = NamedStyle(name="sum_total")
    style.border = Border(bottom=Side(style="double", color="000000"))
    local_styles[style.name] = style

    style = NamedStyle(name="sub_total")
    style.font = Font(bold=True)
    style.border = Border(bottom=Side(style="thin", color="000000"))
    local_styles[style.name] = style

    style = NamedStyle(name="bold")
    style.font = Font(bold=True)
    local_styles[style.name] = style

    style = NamedStyle(name="align_right")
    style.font = Font(bold=True)
    style.border = Border(top=Side(style="thin", color="000000"))
    style.alignment = Alignment(horizontal="right", vertical="center")
    local_styles[style.name] = style

    style = NamedStyle(name="align_left")
    style.font = Font(bold=True)
    style.border = Border(top=Side(style="thin", color="000000"))
    style.alignment = Alignment(horizontal="left", vertical="center")
    local_styles[style.name] = style

    style = NamedStyle(name="align_right_double")
    style.font = Font(bold=True)
    style.border = Border(top=Side(style="double", color="000000"))
    style.alignment = Alignment(horizontal="right", vertical="center")
    local_styles[style.name] = style

    style = NamedStyle(name="align_left_double")
    style.font = Font(bold=True)
    style.border = Border(top=Side(style="double", color="000000"))
    style.alignment = Alignment(horizontal="left", vertical="center")
    local_styles[style.name] = style

    for style in named_styles:
        local_styles[style.name] = style

    for style in local_styles.values():
        wb.add_named_style(style)

    ws.append(headers)

    [ws.append(row) for row in rows]

    if freeze_panes:
        ws.freeze_panes = freeze_panes

    if auto_size:
        autosize_columns(ws)

    if dimensions:
        for key, val in dimensions.get("rows", {}).items():
            ws.row_dimensions[key].height = val
        for key, val in dimensions.get("columns", {}).items():
            ws.column_dimensions[key].width = val

    for cell, style in styles.items():
        ws[cell].style = style

    for cell_range in merge:
        ws.merge_cells(cell_range)

    for cell, format in formats.items():
        ws[cell].number_format = format

    wb.save(fileorbuffer)


def getDownloadableSpreadsheet(
    headers,
    rows,
    styles={},
    merge=[],
    formats={},
    named_styles=[],
    freeze_panes="A2",
    dimensions=None,
    auto_size=True,
):
    buffer = BytesIO()
    create_spreadsheet(
        headers,
        rows,
        buffer,
        styles,
        merge,
        formats,
        named_styles,
        freeze_panes,
        dimensions,
        auto_size,
    )
    return base64.b64encode(buffer.getvalue()).decode()
