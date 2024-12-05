import sys
import csv
from typing import Any


import openpyxl
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, NamedStyle, PatternFill, Side


LIGHT_BLUE_RGB = "D7E4CC"
LIGHT_GREEN_RGB = "D7CCE4"

WHITE_RGB = "FFFFFF"
GREY_RGB = "DFDFDF"


def format_csv(
    input_path: str,
    output_path: str,
    key_main_identifier: str,
    key_version_1: str,
    key_version_2: str,
):
    workbook = openpyxl.Workbook()

    worksheet = workbook.active
    assert worksheet is not None

    def write_row(row, col, data: list[Any], cell_style=None):
        _write_row(worksheet, row, col, data, cell_style=cell_style)

    main_color = AlternatingColor([LIGHT_BLUE_RGB, LIGHT_GREEN_RGB])
    other_color = AlternatingColor([WHITE_RGB, GREY_RGB])

    styles = StyleCache(workbook)

    with open(input_path) as f:
        reader = csv.reader(f)

        input_indexes = {}
        output_indexes = {}

        split_point = None

        header_row = next(reader)

        for idx, header in enumerate(header_row):
            input_indexes[header] = idx

            if split_point is not None:
                # already set
                continue

            # do not use key_main_identifier as split_point
            if key_main_identifier in header:
                continue

            if header.startswith(key_version_2.split("_")[0]):
                # split index should be the first header that
                # contains key_version_2 keyword
                split_point = idx

        if split_point is None:
            # fallback to the last column if could not find a valid header
            split_point = len(header_row)

        # Add our extra "change" header between the two versions we are comparing

        full_headers = list(header_row[:split_point])
        full_headers.append("change")
        full_headers.extend(header_row[split_point:])

        output_indexes = {k: i for i, k in enumerate(full_headers)}

        write_row(0, 0, list(output_indexes.keys()))

        i = 1
        for row in reader:
            if not row:
                continue

            def input_val(field_name: str) -> Any:
                return row[input_indexes[field_name]]

            value_main_identifier = input_val(key_main_identifier)
            value_version_1 = input_val(key_version_1)
            value_version_2 = input_val(key_version_2)

            row_top_border = False
            if main_color.eval_change(value_main_identifier):
                # Add empty row between programs, and top border for the next row
                write_row(i, 0, [""] * len(output_indexes), cell_style=styles.get_style(WHITE_RGB, top_border=True))
                i += 1
                row_top_border = True

            if other_color.eval_change(value_version_2):
                # Add top border for next row to differentiate multiple "other" reports
                row_top_border = True

            write_row(i, 0, row[:split_point], cell_style=styles.get_style(main_color.color(), top_border=row_top_border))

            # Add our extra "change" value between the two reports
            change_status = get_change_status(value_version_1, value_version_2)

            write_row(i, split_point, [change_status], cell_style=styles.get_style(other_color.color(), top_border=row_top_border))

            write_row(i, split_point + 1, row[split_point:], cell_style=styles.get_style(other_color.color(), top_border=row_top_border))

            i += 1

    workbook.save(output_path)


def get_change_status(first_report_cue_id: str, other_report_cue_id: str) -> str:
    if not first_report_cue_id:
        return "ADDED"
    elif not other_report_cue_id:
        return "REMOVED"
    elif other_report_cue_id != first_report_cue_id:
        return "CHANGED"
    else:
        return ""


class AlternatingColor:
    def __init__(self, colors):
        self.colors = colors
        self.index = 0
        self.prev_val = None

    def eval_change(self, new_value) -> bool:
        changed = False
        if self.prev_val is not None and new_value != self.prev_val:
            self.index = (self.index + 1) % len(self.colors)
            changed = True
        self.prev_val = new_value
        return changed

    def color(self):
        return self.colors[self.index]


def _write_row(sheet: Worksheet, row, col, data: list[Any], cell_style=None):
    for i, value in enumerate(data):
        cell = sheet.cell(row=row + 1, column=col + i + 1, value=value)
        if cell_style:
            cell.style = cell_style


class StyleCache:
    def __init__(self, workbook: openpyxl.Workbook):
        self.workbook = workbook
        self.styles = {}

    styles = {}

    def get_style(self, color: str, top_border: bool = False) -> NamedStyle:
        style_name = f"BG: {color}, Top Border: {top_border}"

        if style_name in self.styles:
            return self.styles[style_name]
        else:
            border = None
            if top_border:
                double_border_side = Side(border_style="thin")
                border = Border(top=double_border_side)

            alignment = Alignment(horizontal="general")

            fill = PatternFill(fgColor=color, fill_type=FILL_SOLID)

            style = NamedStyle(style_name)
            if border:
                style.border = border
            style.alignment = alignment
            style.fill = fill

            self.workbook.add_named_style(style)
            self.styles[style_name] = style
            return style


if __name__ == "__main__":
    args = sys.argv[1:]

    if len(args) != 5:
        print("""
            Usage:
                python3 format_report.py <input_path> <output_path> <key_main_identifier> <key_version_1> <key_version_2>

            Where:
                - input_path: path to CSV file
                - output_path: path to XLSX file
                - key_main_identifier: column used to group by the main colours
                - key_version_1: column to compare against key_version_2
                - key_version_2: column to compare against key_version_1

            Example:
                python format_report.py diff_reports/last_reported_vs_prod.csv diff_reports/last_reported_vs_prod.xlsx program_id last_reported_cue_id production_cue_id
        """.replace("            ", ""))
        sys.exit(1)

    format_csv(*args)
