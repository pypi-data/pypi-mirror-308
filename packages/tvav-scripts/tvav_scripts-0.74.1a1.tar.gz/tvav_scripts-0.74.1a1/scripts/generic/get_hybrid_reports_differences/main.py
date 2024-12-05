import base64
import duckdb
import logging
import openpyxl
import os
import pandas as pd
import re
import sys
from bson import ObjectId
from dataclasses import dataclass
from datetime import datetime, timedelta
from dateutil.parser import parse
from dotenv import load_dotenv
from pathlib import Path
from pydantic import BaseModel, Field
from typing import Iterator, Optional

from pymongo import MongoClient
from pymongo.synchronous.database import Database

from scripts.generic.get_hybrid_reports_differences.queries import QueryFactory
from scripts.utils.decorators import ntfy

logger = logging.getLogger("hybrid-diffs")
REPORTAL_URL = "https://kanal-d-reportal.bmat.com/programs/view/"
DATE_FINDER = re.compile(r".*_(\d{4}-\d{2}-\d{2})_.*\.xlsx$")
SCHEDULE_ID_FINDER = re.compile(r'^=HYPERLINK\("https:\/\/kanal-d-reportal\.bmat\.com\/programs\/view\/(.*)#t=.*","Link"\)$')


class Config:
    def __init__(self) -> None:
        load_dotenv()
        reportal_reports_dir = os.getenv("REPORTAL_REPORTS_DIR")
        vericast_reports_dir = os.getenv("VERICAST_REPORTS_DIR")
        self.mongo_uri = os.getenv("MONGO_URI")

        assert reportal_reports_dir is not None
        assert vericast_reports_dir is not None
        assert self.mongo_uri is not None

        if __name__ == "__main__":
            self.parent_dir = Path(__file__).parent
        else:
            self.parent_dir = Path(os.getcwd())

        self.reportal_reports_dir = self.parent_dir / reportal_reports_dir
        self.vericast_reports_dir = self.parent_dir / vericast_reports_dir
        self.script_reports_dir = self.parent_dir / "script_reports"

        self.script_reports_dir.mkdir(parents=True, exist_ok=True)

        reportal_reports_names = {f.name for f in self.reportal_reports_dir.iterdir() if f.name.endswith(".xlsx")}
        vericast_reports_names = {f.name for f in self.vericast_reports_dir.iterdir() if f.name.endswith(".xlsx")}

        same_name_files = reportal_reports_names.intersection(vericast_reports_names)

        reportal_diff_names = reportal_reports_names - same_name_files
        vericast_diff_names = vericast_reports_names - same_name_files

        if (
            len(reportal_diff_names) != 0 or
            len(vericast_diff_names) != 0
        ):
            logger.info(f"{reportal_diff_names=}")
            logger.info(f"{vericast_diff_names=}")
            logger.error("Reports should be named the same in both directories for the comparison to happen")
            sys.exit(-1)


class HybridExcelRow(BaseModel):
    channel: Optional[str] = Field(..., alias="Channel")
    epg_title: Optional[str] = Field(..., alias="EPG Title")
    epg_type: Optional[str] = Field(..., alias="EPG Type")
    link: str = Field(..., alias="Link")
    track: Optional[str] = Field(..., alias="Track")
    artist: Optional[str] = Field(..., alias="Artist")
    label: Optional[str] = Field(..., alias="Label")
    isrc: Optional[str] = Field(..., alias="ISRC")
    iswc: Optional[str] = Field(..., alias="ISWC")
    bmat_id: Optional[str] = Field(..., alias="BmatId")
    album: Optional[str] = Field(..., alias="Album")
    utc_start_time: datetime = Field(..., alias="UTC Start Time")
    utc_end_time: datetime = Field(..., alias="UTC End Time")
    utc_duration_secs: int = Field(..., alias="UTC Duration (s)")


@dataclass()
class LinkTimeshift:
    link: str
    schedule_id: ObjectId
    time_shift: float = 0.0


def get_time_shifts_from_reportal(links: Iterator[str], mongodb: Database) -> Iterator[LinkTimeshift]:
    """Gets the Schedule id from Reportal link and queries Reportal DB to get the time_shift."""

    link_timeshifts: list[LinkTimeshift] = []
    schedule_ids = []
    for link in links:
        if not (match := SCHEDULE_ID_FINDER.match(link)):
            raise RuntimeError(f"Could not get Schedule id from {link=}")

        encoded_schedule_id = match.group(1)

        schedule_id = ObjectId(base64.b64decode(encoded_schedule_id.encode()).decode().replace("Schedule:", ""))
        schedule_ids.append(schedule_id)

        link_timeshifts.append(LinkTimeshift(link=link, schedule_id=schedule_id))

    schedule_ids_time_shifts = {
        sch["_id"]: sch["time_shift"]
        for sch in mongodb.schedule.find({"_id": {"$in": schedule_ids}}, {"time_shift": 1})
    }

    for link_timeshift in link_timeshifts:
        link_timeshift.time_shift = schedule_ids_time_shifts[link_timeshift.schedule_id]
        yield link_timeshift


@ntfy()
def read_hybrid_excel(excel_file_path: str, mongodb: Database) -> pd.DataFrame:
    """Parses useful cols from Hybrid Excel report into a pandas DataFrame."""

    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb[wb.sheetnames[0]]

    parsed_rows = []

    rows = sheet.iter_rows()
    header = [c.value for c in next(rows)]
    for row in rows:
        row_data = dict.fromkeys(header)

        for col_idx, cell in enumerate(row):
            if cell.hyperlink:
                row_data[header[col_idx]] = cell.hyperlink.target
            else:
                row_data[header[col_idx]] = cell.value

        try:
            row_data_obj = HybridExcelRow.model_validate(row_data)
        except Exception as e:
            utc_duration_secs = row_data["UTC Duration (s)"]
            if isinstance(utc_duration_secs, str) and "SUM" in utc_duration_secs:
                continue
            raise e

        parsed_rows.append(row_data_obj.model_dump(by_alias=True))

    # remove time_shift in Reportal Reports UTC Start Time and UTC End Time
    if (
        (link := parsed_rows[0]["Link"]) and
        isinstance(link, str) and
        REPORTAL_URL in link
    ):
        time_shift_removed_parsed_rows = []
        for link_timeshift in get_time_shifts_from_reportal(
            set(row["Link"] for row in parsed_rows),  # type: ignore
            mongodb
        ):
            for row in filter(lambda r: r["Link"] == link_timeshift.link, parsed_rows):
                row["UTC Start Time"] -= timedelta(seconds=link_timeshift.time_shift)
                row["UTC End Time"] -= timedelta(seconds=link_timeshift.time_shift)

                time_shift_removed_parsed_rows.append(row)

        if len(time_shift_removed_parsed_rows) != len(parsed_rows):
            raise RuntimeError("Lengths do not match")

        parsed_rows = time_shift_removed_parsed_rows

    return pd.DataFrame(parsed_rows)


@ntfy()
def ingest_local_db(
    config: Config,
    con: duckdb.DuckDBPyConnection,
    mongodb: Database
):
    """Ingests local DB with Hybrid reports information from Reportal and Vericast."""

    for reports_dir in [config.reportal_reports_dir, config.vericast_reports_dir]:
        for hybrid_report in (
            f
            for f in
            reports_dir.iterdir()
            if f.name.endswith(".xlsx")
        ):
            report_date_match = DATE_FINDER.match(hybrid_report.name)
            assert report_date_match is not None, hybrid_report.name
            report_date = parse(report_date_match.group(1))

            table_name = f"{hybrid_report.parent.name}_{report_date.month}"

            if con.sql(f"select 1 from information_schema.tables where table_name = '{table_name}'").fetchone():
                logger.info(f"Already exists table {table_name}")
                continue

            logger.info(f"Inserting {hybrid_report.parent.name}/{hybrid_report.name} into {table_name}")

            # NOTE: ignore linter, it is being used by duckdb using monkey-patching
            data = read_hybrid_excel(str(hybrid_report.absolute()), mongodb)

            con.sql(f"CREATE OR REPLACE SEQUENCE serial_{table_name}")
            con.sql(f"""
                CREATE OR REPLACE TABLE {table_name} (
                    id INTEGER PRIMARY KEY DEFAULT nextval('serial_{table_name}'),
                    Channel TEXT,
                    "EPG Title" TEXT,
                    "EPG Type" TEXT,
                    Link TEXT,
                    Track TEXT,
                    Artist TEXT,
                    Label TEXT,
                    ISRC TEXT,
                    ISWC TEXT,
                    BmatId TEXT,
                    Album TEXT,
                    "UTC Start Time" DATETIME,
                    "UTC End Time" DATETIME,
                    "UTC Duration (s)" INT
                )
            """)
            con.sql(f"""
                INSERT INTO {table_name} (
                    Channel,
                    "EPG Title",
                    "EPG Type",
                    Link,
                    Track,
                    Artist,
                    Label,
                    ISRC,
                    ISWC,
                    BmatId,
                    Album,
                    "UTC Start Time",
                    "UTC End Time",
                    "UTC Duration (s)"
                )
                    SELECT * FROM data
            """)


def generate_reports(config: Config, con: duckdb.DuckDBPyConnection):
    """Queries local DB and writtes the output into XLSX report files."""

    for tbl_index in [4, 5, 6]:
        for query in QueryFactory(report_month=tbl_index):
            report_file = config.script_reports_dir / f"{query.report_filename}.xlsx"
            logger.info(f"Writting {report_file.name=}")

            res = con.sql(query.sql_str)

            wb = openpyxl.Workbook()
            ws = wb.active
            assert ws is not None

            ws.append(query.col_names)
            while row := res.fetchone():
                ws.append(row)

            wb.save(report_file)
    

@ntfy()
def run():
    config = Config()
    con = duckdb.connect(config.parent_dir / "hybrids_diff.duckdb")
    mongodb = MongoClient(host=config.mongo_uri).get_default_database()

    ingest_local_db(config, con, mongodb)

    generate_reports(config, con)

if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s:%(levelname)s:%(name)s: %(message)s'
    )
    run()
