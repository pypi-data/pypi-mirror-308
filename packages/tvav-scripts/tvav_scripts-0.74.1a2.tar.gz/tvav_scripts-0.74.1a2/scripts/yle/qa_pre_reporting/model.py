from datetime import datetime
from itertools import filterfalse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import Field, BaseSettings, BaseModel
from typing import List, Any, Optional, Union

from reportal_model import AvWork, Schedule, Cue

from scripts.yle.qa_pre_reporting.utils import (
    contributor_with_role_exists_in_sv, DEFAULT_START_TIME,
    get_reportal_cuesheet_url, check_music_work_contributors,
    CONTRIBUTOR_AUTHOR_ROLES, CONTRIBUTOR_INTERPRETER_ROLES,
)


class QAPreReportingScriptConfig(BaseSettings):
    """To establish DB connection, generate report document and report filename"""
    mongo_credentials: str = Field(..., env="mongo_credentials")
    mongo_db: str = Field(..., env="mongo_db")
    username: str = Field("bmat_processor", env="username")
    start_time: datetime = DEFAULT_START_TIME
    report_date: datetime = Field(..., env="report_date")

    def get_mongodb_connection_string(self) -> str:
        """
        Uses the mongo credentials and db to generate the connection string.
        :return: The mongodb connection string.
        """
        return (
            "mongodb+srv://{}@bmat-tvav-prod.yq6o5.mongodb.net/{}?retryWrites=true&w=majority".format(
                self.mongo_credentials,
                self.mongo_db
            )
        )


class QACuesheetMissingContributorRow(BaseModel):
    cue_index: int = Field(..., alias="Cue #")
    single_view_id: Optional[str] = Field(..., alias="Single View ID")
    source: Optional[str] = Field(..., alias="Source")
    service_id: Optional[str] = Field(..., alias="Service id")
    missing_in_sv: Optional[Union[bool, str]] = Field(..., alias="Missing in SV too (False = fixed with re-enrichment)")
    reportal_url: str = Field(..., alias="Reportal URLs")

    class Config:
        allow_population_by_field_name = True


class QACuesheetsMissingContributor(BaseModel):
    """Class with stats for cuesheets with missing contributor."""
    total: int = Field(0, alias="Total")
    rows: list[QACuesheetMissingContributorRow] = Field(default_factory=list)

    def add_cuesheet(self, av_work: AvWork, cue_index: int, roles_to_check: list[str]) -> None:
        """
        Increases the total number of cuesheets in the report + adds the cuesheet URL to the report.
        :param av_work: The AvWork object
        :param cue_index: Cue index as an int starting from 0
        :param roles_to_check: list[str] of roles to check in SV API
        """
        self.total += 1
        missing_in_sv = None

        cue = av_work.cuesheet.cues[cue_index - 1]
        single_view_id = cue.music_work.work_ids.get("single_view_id")
        source = None
        service_id = None

        if cue.cue_identifiers:
            source = cue.cue_identifiers.source
            service_id = cue.cue_identifiers.service_id

        if single_view_id or (source and service_id):
            missing_in_sv = not contributor_with_role_exists_in_sv(single_view_id, source, service_id, roles_to_check)
        else:
            missing_in_sv = "Manual edition"

        self.rows.append(QACuesheetMissingContributorRow(
            cue_index=cue_index,
            single_view_id=single_view_id,
            source=source,
            service_id=service_id,
            missing_in_sv=missing_in_sv,
            reportal_url=get_reportal_cuesheet_url(av_work),
        ))


class QAReportedIdentifiedMusic(BaseModel):
    """Class with stats for music works."""
    total: int = Field(0, alias="Total")
    total_duration: float = Field(0, alias="Total duration")

    def add_cue(self, cue: Cue):
        """
        Adds the cue duration to the report and increments the total number of identified cues added.
        :param cue: the identified cue to add to the QA report
        :raise ValueError: if duration is missing
        """
        self.total += 1
        try:
            self.total_duration += cue.duration
        except TypeError:
            raise ValueError("Cue:%s missing duration" % str(cue.id))


class QAError(BaseModel):
    """Class with all error info for the QA Report"""
    reportal_url: str
    music_work_title: str = None
    error_description: str


class QAReportChannel(BaseModel):
    """
    Class with stats for a Channel:

    - sheet_title: to be used as the Excel sheet title
    - total_reported_cuesheets: sum of all AvWorks reported
    - total_reported_broadcasts: sum of all Schedules reported
    - reported_identified_music: Identified music stats
    - cuesheets_missing_interpreter: AvWorks with missing interpreter (performer) stats
    - cuesheets_missing_author: AvWorks with missing composer stats
    - errors: Error stats
    """
    sheet_title: str
    total_reported_cuesheets: int = Field(0, alias="Total reported Cuesheets")
    total_reported_broadcasts: int = Field(0, alias="Total reported Broadcasts")
    reported_identified_music: QAReportedIdentifiedMusic = Field(QAReportedIdentifiedMusic(), alias="Reported identified music")
    cuesheets_missing_interpreter: QACuesheetsMissingContributor = Field(QACuesheetsMissingContributor(), alias="Cuesheets missing interpreter (performer)")
    cuesheets_missing_author: QACuesheetsMissingContributor = Field(QACuesheetsMissingContributor(), alias="Cuesheets missing author (composer)")
    errors: list[QAError] = Field(default_factory=list, alias="Errors")

    def add_broadcast(self, schedule: Schedule):
        """
        Increases the total reported broadcasts by 1 and adds the linked AvWork if any
        :param schedule: Broadcast to be added to the report
        :return:
        """
        self.total_reported_broadcasts += 1

        if schedule.av_work:
            self._add_cuesheet(schedule.av_work)

    def _add_cuesheet(self, av_work: AvWork):
        """
        Increases the total reported cuesheets by 1, registers the MusicWorks and checks cuesheet contributors.

        Will also populate the errors field with all error instances.
        :param av_work: Cuesheet to add to the report
        :return:
        """
        self.total_reported_cuesheets += 1

        # quick return if no cuesheet
        if av_work.cuesheet is None:
            return

        for id_cue in filterfalse(lambda cue: not cue.music_work, av_work.cuesheet.cues):
            music_work = id_cue.music_work
            try:
                self.reported_identified_music.add_cue(id_cue)
                has_author, has_interpreter = check_music_work_contributors(music_work)
                if not has_author:
                    self.cuesheets_missing_author.add_cuesheet(av_work, id_cue.cue_index, CONTRIBUTOR_AUTHOR_ROLES)
                if not has_interpreter:
                    self.cuesheets_missing_interpreter.add_cuesheet(av_work, id_cue.cue_index, CONTRIBUTOR_INTERPRETER_ROLES)
            except ValueError as e:
                self.errors.append(QAError(
                    reportal_url=get_reportal_cuesheet_url(av_work),
                    music_work_title=music_work.title,
                    error_description=str(e),
                ))


class QAReportGenerator(BaseModel):
    """Class to generate the QA report."""
    wb: Optional[Workbook] = None
    filename: Optional[str] = None
    summary: Optional[Worksheet] = None
    summary_row_index: int = 1

    class Config:
        arbitrary_types_allowed = True

    def _add_summary_headers(self):
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')

        if self.summary is None:
            raise RuntimeError("Summary sheet is missing")

        self.summary["A1"] = "Cue #"
        self.summary["B1"] = "Single View ID"
        self.summary["C1"] = "Source"
        self.summary["D1"] = "Service id"
        self.summary["E1"] = "Missing in SV too (False = fixed with re-enrichment)"
        self.summary["F1"] = "Reportal URLs"
        
        for row in self.summary:
            for cell in row:
                if cell.value:
                    cell.font = header_font
                    cell.alignment = header_alignment

    @staticmethod
    def _add_headers(ws):
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')

        # First row
        ws["A1"] = "Total reported cuesheets"
        ws["B1"] = "Total reported broadcasts"
        ws["C1"] = "Reported identified music"
        ws["E1"] = "Cuesheets missing interpreter (performer)"
        ws["L1"] = "Cuesheets missing author (composer)"
        ws["S1"] = "Errors"
        # Second row
        ws["C2"] = "Total"
        ws["D2"] = "Total duration"
        ws["E2"] = "Total"
        ws["F2"] = "Cue #"
        ws["G2"] = "Single View ID"
        ws["H2"] = "Source"
        ws["I2"] = "Service id"
        ws["J2"] = "Missing in SV too (False = fixed with re-enrichment)"
        ws["K2"] = "Reportal URLs"
        ws["L2"] = "Total"
        ws["M2"] = "Cue #"
        ws["N2"] = "Single View ID"
        ws["O2"] = "Source"
        ws["P2"] = "Service id"
        ws["Q2"] = "Missing in SV too (False = fixed with re-enrichment)"
        ws["R2"] = "Reportal URLs"
        ws["S2"] = "Reportal URLs"
        ws["T2"] = "Music Work Title"
        ws["U2"] = "error_description"

        # Update font and alignment
        for row in ws:
            for cell in row:
                if cell.value:
                    cell.font = header_font
                    cell.alignment = header_alignment

        ws.merge_cells("C1:D1")
        ws.merge_cells("E1:K1")
        ws.merge_cells("L1:R1")
        ws.merge_cells("S1:U1")

    def _add_values(self, ws, qa_report: QAReportChannel):
        ws["A2"] = qa_report.total_reported_cuesheets
        ws["B2"] = qa_report.total_reported_broadcasts
        ws["C3"] = qa_report.reported_identified_music.total
        ws["D3"] = qa_report.reported_identified_music.total_duration
        ws["E3"] = qa_report.cuesheets_missing_interpreter.total
        ws["L3"] = qa_report.cuesheets_missing_author.total

        def _add_contributor_values(ws, start_col: str, rows: list[QACuesheetMissingContributorRow]):
            start_col_ord = ord(start_col)
            for row_index, row_value in enumerate(rows, 3):
                ws[f"{start_col}{row_index}"] = row_value.cue_index
                ws[f"{chr(start_col_ord + 1)}{row_index}"] = row_value.single_view_id
                ws[f"{chr(start_col_ord + 2)}{row_index}"] = row_value.source
                ws[f"{chr(start_col_ord + 3)}{row_index}"] = row_value.service_id
                ws[f"{chr(start_col_ord + 4)}{row_index}"] = row_value.missing_in_sv
                ws[f"{chr(start_col_ord + 5)}{row_index}"] = row_value.reportal_url

                if self.summary is None:
                    raise RuntimeError("Summary sheet is missing")

                # update summary sheet
                self.summary[f"A{self.summary_row_index}"] = str(row_value.cue_index)
                self.summary[f"B{self.summary_row_index}"] = getattr(row_value, "single_view_id", "")
                self.summary[f"C{self.summary_row_index}"] = getattr(row_value, "source", "")
                self.summary[f"D{self.summary_row_index}"] = getattr(row_value, "service_id", "")
                self.summary[f"E{self.summary_row_index}"] = getattr(row_value, "missing_in_sv", "")
                self.summary[f"F{self.summary_row_index}"] = getattr(row_value, "reportal_url", "")
                self.summary_row_index += 1

        _add_contributor_values(ws, "F", qa_report.cuesheets_missing_interpreter.rows)
        _add_contributor_values(ws, "M", qa_report.cuesheets_missing_author.rows)

        # leave 1 empty row of separation
        self.summary_row_index += 1

        for row_index, row_value in enumerate(qa_report.errors, 3):
            ws[f"S{row_index}"] = row_value.reportal_url
            ws[f"T{row_index}"] = row_value.music_work_title
            ws[f"U{row_index}"] = row_value.error_description

    @staticmethod
    def new_channel_report(channel_name: str):
        return QAReportChannel(sheet_title=channel_name)

    def add_channel_report(self, channel_report: QAReportChannel):
        """
        Generates the QA Report in Excel format with the specified filename.
        :param channel_report:
        :return:
        """
        ws = self.wb.create_sheet(title=channel_report.sheet_title)

        self._add_headers(ws)
        self._add_values(ws, channel_report)

        self.wb.save(self.filename)

    def create_report(self, filename: str):
        """Creates the Workbook and deletes the first sheet"""
        self.filename = filename
        self.wb = Workbook()

        self.summary = self.wb.active
        self.summary.title = "Summary"
        self._add_summary_headers()
        self.summary_row_index += 1
