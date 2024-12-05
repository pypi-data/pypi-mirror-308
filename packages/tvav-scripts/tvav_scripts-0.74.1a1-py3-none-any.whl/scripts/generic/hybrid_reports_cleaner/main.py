import openpyxl
import os
from alphabet_detector import AlphabetDetector
from pathlib import Path
from tqdm import tqdm

if __name__ == "__main__":
    parent_dir = Path(__file__).parent
else:
    parent_dir = Path(os.getcwd())

originals_dir = parent_dir / "originals"
cleaned_dir = parent_dir / "cleaned"


def clean_hybrid_report(original_file: Path) -> None:
    """Generates a clean hybrid report from original_file into cleaned_dir."""

    wb = openpyxl.load_workbook(original_file)
    wb_clean = openpyxl.Workbook()

    sheet = wb[wb.sheetnames[0]]
    sheet_clean = wb_clean.active
    assert sheet_clean is not None

    rows = sheet.iter_rows(values_only=True)

    header = next(rows)
    sheet_clean.append(header)

    ad = AlphabetDetector()
    for row_data in tqdm(rows, leave=False):
        # only latin chars
        if not all([
            ad.only_alphabet_chars(v, "LATIN")
            for v in row_data
            if isinstance(v, str)
        ]):
            continue

        row = dict(zip(header, row_data))
        
        # no missing epg info
        skip_row = False
        for k, v in row.items():
            if (
                isinstance(k, str) and
                k.startswith("EPG") and
                v in [None, "", " "]
            ):
                skip_row = True
                break

        if skip_row:
            continue

        # no duration less than 3
        duration = row.get("Duration")
        if duration is None:
            continue
        if not isinstance(duration, str):
            raise TypeError(duration)

        duration_seconds = sum([int(v) for v in duration.split(":")])
        if duration_seconds < 3:  # type: ignore
            continue

        sheet_clean.append(row_data)

    wb_clean.save(cleaned_dir / original_file.name)


if __name__ == "__main__":
    [
        clean_hybrid_report(report_file)
        for report_file in tqdm(originals_dir.iterdir(), total=len([_ for _ in originals_dir.iterdir()]))
    ]
