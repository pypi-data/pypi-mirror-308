from barcode.writer import ImageWriter
from PIL import ImageDraw, ImageFont
from qrcode import QRCode, constants
from collections import defaultdict
from dataclasses import dataclass
from gkeepapi import Keep, _node
from mailmerge import MailMerge
import plotly.graph_objs as go
from xlutils.copy import copy
from datetime import datetime
from PIL.Image import Image
import PIL.Image as mImage
from typing import Any
from os import path
import xlsxwriter
import openpyxl
import barcode
import xlrd
import io


from pih import A
from DocsService.const import *
from pih.collections import (
    Note,
    Result,
    FullName,
    FieldItem,
    LoginPasswordPair,
    TimeSeriesStatistics,
    TimeTrackingResultByDate,
    TimeTrackingResultByPerson,
    CTIndicationsValueContainer,
    TimeTrackingResultByDivision,
)
from pih.tools import j, nn, ne, n, e, nl
from pih.consts.errors import IncorrectInputFile


@dataclass
class BarCodeData:
    value: str
    name: str


class DocumentApi:
    
    def __init__(self, test: bool = False):
        if not test and A.S.get(A.CT_S.USE_GOOGLE_KEEP):
            self.keep = Keep()
            self.keep.login(
                A.D_V_E.value(GKEEP_USER_LINK.LOGIN),
                A.D_V_E.value(GKEEP_USER_LINK.PASSWORD),
            )
            A.O.good("google keep inialized")
        self.workbook_map: dict = {}

    def to_hh_mm(self, value: int) -> str:
        min, _ = divmod(value, 60)
        hour, min = divmod(min, 60)
        return "%d:%02d:00" % (hour, min)

    def create_note(
        self,
        note: Note,
        label_list: list[str] | None = None,
        color: str | None = None,
        remove_before: bool = False,
    ) -> int:
        gnote: _node.Node | None = None
        if remove_before:
            for gnote in self.keep.find(func=lambda item: item.title == note.title):
                gnote.trash()
        gnote = self.keep.createNote(note.title, note.text)
        if n(label_list):
            gnote.labels.add(self.keep.findLabel("Мобильные заметки"))
        else:
            A.D.every(
                lambda item: gnote.labels.add(self.keep.findLabel(item)), label_list
            )
        gnote.pinned = True
        gnote.color = A.D.get_by_value(_node.ColorValue, color) or _node.ColorValue.Blue
        self.keep.sync()
        return gnote.id

    def get_note(self, id: str, force_update: bool = False) -> Note:
        if force_update:
            self.drop_note_cache()
        gnote = self.keep.get(id)
        images: list[str] = []
        if ne(gnote.images):
            for image in gnote.images:
                images.append(self.keep.getMediaLink(image))
        return Note(gnote.title, gnote.text, id, images)

    def get_notes_by_label(self, value: str, force_update: bool = False) -> list[Note]:
        if force_update:
            self.drop_note_cache()
        result: list[Note] = []
        for gnote in self.keep.find(labels=[self.keep.findLabel(value)]):
            images: list[str] = []
            if ne(gnote.images):
                for image in gnote.images:
                    images.append(self.keep.getMediaLink(image))
            result.append(Note(gnote.title, gnote.text, gnote.id, images))
        return result

    def create_statistics_chart(self, name: str) -> bool:
        if name == A.D.get(A.CT_STAT.Types.CHILLER_FILTER):
            statistics: TimeSeriesStatistics = A.D_STAT.by_name(name)
            value_y: list[float] = A.D.map(A.D.seconds_to_days, statistics.distance)
            value_x: list[datetime] = statistics.values
            bax_x = [
                f"{index + 1}: {A.D.to_date_string(datetime_start.isoformat())}"
                for index, datetime_start in enumerate(value_x)
            ][:-1]
            bins: list[int] = [0, 7, 10, 14, 100]
            labels: list[str] = ["Плохо", "Нормально", "Хорошо", "Отлично"]
            for index, item in enumerate(bins):
                if index == len(bins) - 1:
                    break
                if index == 0:
                    labels[index] = labels[index] + f": <{bins[index + 1]}"
                elif index == len(bins) - 2:
                    labels[index] = labels[index] + f": >{item}"
                else:
                    labels[index] = labels[index] + f": {item} - {bins[index + 1]}"
            colors: list[str] = ["lightcoral", "lightgreen", "darkgreen", "orange"]
            result: dict[str, tuple[str, str, list[float]]] = {}
            for index, label in enumerate(labels):
                result[label] = (label, colors[index], [])
                min_y: int = bins[index]
                max_y: int = bins[index + 1]
                for y in value_y:
                    if y > min_y and y < max_y:
                        result[label][2].append(y)
                    else:
                        result[label][2].append(0)
            bars = []
            for index, label in enumerate(result):
                item: tuple[str, str, list[float]] = result[label]
                bars.append(
                    go.Bar(x=bax_x, y=item[2], name=item[0], marker={"color": item[1]})
                )
            go.FigureWidget(
                data=bars,
                layout=go.Layout(
                    barmode="stack",
                    title="Время выработки фильтров чиллера МРТ",
                    yaxis_title="Дни",
                    xaxis_title="номер: дата начала выработки",
                ),
            ).write_image(A.PTH_STAT.get_file_path(name))
            return True
        ct_statistics_type_names: list[str] = [
            A.D.get(A.CT_STAT.Types.CT_DAY),
            A.D.get(A.CT_STAT.Types.CT_WEEK),
            A.D.get(A.CT_STAT.Types.CT_MONTH),
        ]
        work_ct_statistics_type_names: list[str] | None = None
        if name in ct_statistics_type_names:
            work_ct_statistics_type_names = [name]
        elif name == A.D.get(A.CT_STAT.Types.CT):
            work_ct_statistics_type_names = ct_statistics_type_names
        if A.D.is_not_none(work_ct_statistics_type_names):
            for work_ct_statistics_type_name in work_ct_statistics_type_names:
                type_index: int = -1
                for index, item in enumerate(ct_statistics_type_names):
                    if item == work_ct_statistics_type_name:
                        type_index = index
                        break
                if type_index != -1:
                    count: int = 24
                    if type_index == 1:
                        count = 7 * 24
                    elif type_index == 2:
                        count = 30 * 24
                    data: list[CTIndicationsValueContainer] = (
                        A.R_IND.last_ct_value_containers(False, count).data
                    )
                    figure = go.Figure()
                    x: list[datetime] = A.D.map(lambda item: item.timestamp, data)
                    figure.add_trace(
                        go.Scatter(
                            x=x,
                            mode="lines",
                            y=A.D.map(lambda item: item.temperature, data),
                            name="Температура",
                        )
                    )
                    figure.add_trace(
                        go.Scatter(
                            x=x,
                            y=A.D.map(lambda item: item.humidity, data),
                            mode="lines",
                            name="Влажность",
                        )
                    )
                    figure.update_layout(
                        legend_orientation="h",
                        xaxis_title="Время",
                        yaxis_title=f"Влажность (%) и Температура ({A.CT_V.TEMPERATURE_SYMBOL})",
                        margin=dict(l=0, r=0, t=0, b=0),
                    )
                    figure.write_image(
                        A.PTH_STAT.get_file_path(work_ct_statistics_type_name)
                    )
            return True
        return False

    def save_xlsx(
        self,
        title: str,
        fields: list[dict],
        data: list[dict[str, Any]],
        path: str,
    ) -> bool:
        try:
            field_list: list[FieldItem] = []
            for field_data in fields:
                for _, field_data in field_data.items():
                    field_list.append(
                        A.D.fill_data_from_source(FieldItem(), field_data)
                    )
            workbook = xlsxwriter.Workbook(path)
            title_format = workbook.add_format(
                {
                    "bold": 1,
                    "border": 1,
                    "border_color": "#ffffff",
                    "align": "center",
                    "valign": "vcenter",
                    "fg_color": "#66ff99",
                }
            )
            worksheet = workbook.add_worksheet()
            worksheet.set_row(0, 20)
            worksheet.set_column("A:A", 40)
            worksheet.set_column("B:B", 40)
            worksheet.set_column("C:E", 40)
            worksheet.set_column("F:F", 40)
            worksheet.set_column("G:G", 40)
            worksheet.set_column("H:H", 40)
            field: FieldItem | None = None
            for column, field in enumerate(field_list):
                worksheet.write_string(0, column, field.caption, title_format)
            for column, field in enumerate(field_list):
                for row, data_item in enumerate(data):
                    value: Any | None = data_item[field.name]
                    worksheet.write_string(
                        row + 1, column, "" if e(value) else str(value)
                    )
            workbook.close()
            return True
        except Exception as _:
            return False

    def save_time_tracking_report(
        self,
        time_traking_result: Result[list[TimeTrackingResultByDivision]],
        path: str,
        plain_format: bool = False,
    ) -> bool:
        try:
            result_data_list = sorted(
                time_traking_result.data,
                key=lambda item: item.name if item.name is not None else "",
            )
            workbook = xlsxwriter.Workbook(path)
            if plain_format:
                worksheet = workbook.add_worksheet("Учет рабочего времени")
                worksheet.set_row(0, 20)
                worksheet.set_column("A:A", 20)
                worksheet.set_column("B:B", 40)
                worksheet.set_column("C:E", 15)
                worksheet.set_column("F:F", 20)
                worksheet.set_column("G:G", 15)
                worksheet.set_column("H:H", 20)
                person_map: list[TimeTrackingResultByPerson] = []
                division_person_map: dict[str, str] = defaultdict(list)
                for division_item in result_data_list:
                    if division_item.name is not None:
                        for person_item in division_item.list:
                            division_person_map[person_item.tab_number] = (
                                division_item.name
                            )
                            person_map.append(person_item)
                index: int = 0
                for item in person_map:
                    for date_item in item.list:
                        date: TimeTrackingResultByDate = date_item
                        worksheet.write_string(
                            index, 0, division_person_map[item.tab_number]
                        )
                        worksheet.write_string(index, 1, item.full_name)
                        worksheet.write_string(index, 2, date.date)
                        worksheet.write_string(index, 3, date.enter_time or "")
                        worksheet.write_string(index, 4, date.exit_time or "")
                        worksheet.write_string(
                            index,
                            5,
                            (
                                ""
                                if date.duration is None
                                else self.to_hh_mm(date.duration)
                            ),
                        )
                        index += 1
            else:
                title_format = workbook.add_format(
                    {
                        "bold": 1,
                        "border": 1,
                        "border_color": "#ffffff",
                        "align": "center",
                        "valign": "vcenter",
                        "fg_color": "#66ff99",
                    }
                )
                tab_number_format = workbook.add_format(
                    {
                        "bold": 1,
                        "border": 0,
                        "align": "center",
                        "valign": "vcenter",
                        "fg_color": "#99ccff",
                    }
                )
                duration_format = workbook.add_format(
                    {
                        "bold": 1,
                        "border": 0,
                        "align": "center",
                        "valign": "vcenter",
                        "fg_color": "#99ccff",
                    }
                )
                person_format = workbook.add_format(
                    {"bold": 1, "border": 0, "fg_color": "#99ccff"}
                )
                full_duration_format = workbook.add_format(
                    {
                        "bold": 1,
                        "border": 0,
                        "align": "right",
                        "valign": "vcenter",
                        "fg_color": "#0066cc",
                    }
                )
                for division_item in result_data_list:
                    if division_item.name is not None:
                        division_item.name = division_item.name or "Без названия"
                        worksheet = workbook.add_worksheet(
                            division_item.name
                            if len(division_item.name) <= EXCEL_TITLE_MAX_LENGTH
                            else division_item.name[0:EXCEL_TITLE_MAX_LENGTH]
                        )
                        worksheet.set_row(0, 20)
                        for item_index, title_item in enumerate(
                            time_traking_result.fields.list
                        ):
                            field_item: FieldItem = title_item
                            worksheet.write_string(
                                0, item_index, field_item.caption, title_format
                            )
                        worksheet.write_string(
                            0, item_index + 1, "Всего часов", title_format
                        )
                        worksheet.write_string(
                            0, item_index + 2, "Часов по табелю", title_format
                        )
                        index: int = 1
                        worksheet.set_column("A:A", 40)
                        worksheet.set_column("B:B", 20)
                        worksheet.set_column("C:E", 15)
                        worksheet.set_column("F:F", 20)
                        worksheet.set_column("G:G", 15)
                        worksheet.set_column("H:H", 20)
                        for person_item in division_item.list:
                            worksheet.write_string(
                                index, 0, person_item.full_name, person_format
                            )
                            worksheet.write_string(
                                index, 1, person_item.tab_number, tab_number_format
                            )
                            length = len(person_item.list)
                            table_data: list = []
                            for date_item in person_item.list:
                                date: TimeTrackingResultByDate = date_item
                                table_data.append(
                                    [
                                        date.date,
                                        date.enter_time,
                                        date.exit_time,
                                        self.to_hh_mm(date.duration),
                                    ]
                                )
                            worksheet.add_table(
                                index,
                                2,
                                index + length - 1,
                                5,
                                {
                                    "autofilter": False,
                                    "data": table_data,
                                    "header_row": False,
                                },
                            )
                            worksheet.merge_range(
                                index + length,
                                2,
                                index + length,
                                4,
                                "Общая продолжительность:",
                                full_duration_format,
                            )
                            worksheet.write_string(
                                index + length,
                                5,
                                self.to_hh_mm(person_item.duration),
                                duration_format,
                            )
                            index += length + 2
            workbook.close()
            return True
        except:
            return False

    def create_user_document(
        self,
        path: str,
        date_now_string: str,
        web_site_name: str,
        web_site: str,
        email_address: str,
        full_name: FullName,
        tab_number: str,
        pc: LoginPasswordPair,
        polibase: LoginPasswordPair,
        email: LoginPasswordPair,
    ) -> bool:
        self.create(
            path,
            date_now_string,
            web_site_name,
            web_site,
            email_address,
            A.D.fullname_to_string(full_name),
            tab_number,
            pc,
            polibase,
            email,
        )
        return True

    def close_inventory_report(self, report_file_path: str) -> bool:
        self.close_inventory_report(report_file_path)
        return True

    def create_barcode_for_polibase_person(self, pin: int, test: bool = True) -> bool:
        settings: dict[str, int | float] = {
            "module_width": 0.2,
            "module_height": 3.5,
            "font_size": 5,
            "text_distance": 2,
            "quiet_zone": 0,
        }
        return self.create_barcode(
            BarCodeData(str(pin), A.CT_P.BARCODE.get_file_name(pin)),
            A.PTH_P.person_folder(pin, test),
            A.CT_P.BARCODE.PERSON.IMAGE_FORMAT,
            A.CT_P.BARCODE.ACTUAL_FORMAT,
            settings,
        )

    def create_qr_code(
        self, data: str, title: str, file_path: str, font_size: int | None = None
    ) -> bool:
        def get_text_size(title: str, font_name: str, font_size: int = 36):
            backgroundRGB = (255, 255, 255)
            img: Image = mImage.new("RGB", (1, 1), backgroundRGB)
            draw: ImageDraw.ImageDraw = ImageDraw.Draw(img)
            font: ImageFont.ImageFont = ImageFont.truetype(font_name, font_size)
            title_width, title_height = draw.textsize(title, font=font)
            return title_width, title_height

        def _create_qr_code_image_with_title(
            path: str,
            data: str,
            title: str,
            font_path: str | None = None,
            font_size: int | None = None,
        ) -> None:
            font_path = font_path or A.PTH.FONTS.get("ubuntu")
            font_size = font_size or 58
            border: int = 15
            vertical_space: int = 15
            qr: QRCode = QRCode(
                error_correction=constants.ERROR_CORRECT_L,
                border=2,
            )
            data = "".join(data)
            qr.add_data(data)
            qr.make()
            img: Image = qr.make_image().convert("RGB")
            textSize = get_text_size(title, font_path, font_size)
            width: int = max((img.size[0], textSize[0])) + border * 2
            height: int = img.size[1] + textSize[1] + border * 2 + vertical_space
            result_image: Image = mImage.new(
                "RGBA", (width, height), (255, 255, 255, 255)
            )
            textTopLeft: tuple[int, int] = ((width - textSize[0]) // 2, border)
            result_image.paste(
                img, ((width - img.size[0]) // 2, textSize[1] + border + vertical_space)
            )
            result: ImageDraw.ImageDraw = ImageDraw.Draw(result_image)
            font: ImageFont.ImageFont = ImageFont.truetype(font_path, font_size)
            result.text(textTopLeft, title, (0, 0, 0), align="center", font=font)
            result_image.save(path)

        _create_qr_code_image_with_title(file_path, data, title, font_size=font_size)
        return True

    def create_barcode(
        self,
        data: BarCodeData,
        result_directory: str,
        image_format: str | None = None,
        barcode_type: str | None = None,
        settings: dict | None = None,
    ) -> bool:
        return self.create_barcodes(
            [data], result_directory, image_format, barcode_type, settings
        )

    def create_barcodes(
        self,
        data: list[BarCodeData],
        result_directory: str,
        image_format: str | None = None,
        barcode_type: str | None = None,
        settings: dict | None = None,
    ) -> bool:
        image_format = image_format or "png"
        barcode_type = barcode_type or "code128"
        settings = settings or {
            "module_width": 0.35,
            "module_height": 10,
            "font_size": 12,
            "text_distance": 5,
            "quiet_zone": 1,
        }
        bar_class = barcode.get_barcode_class(barcode_type)
        writer = ImageWriter(format=image_format)
        for data_item in data:
            barcode_creator = bar_class(data_item.value, writer)
            barcode_creator.save(path.join(result_directory, data_item.name), settings)
        return True

    def create_inventory_barcodes(
        self, report_file_path: str, result_directory: str
    ) -> bool:
        result: list[dict] = self.inventory_report(report_file_path)
        if result is not None:
            barcode_data: list[dict[str, str]] = []
            for index, data_item in enumerate(result):
                barcode_data.append(
                    BarCodeData(
                        data_item[A.CT_FNC.INVENTORY_NUMBER],
                        f"{index + 1} {A.PTH.replace_prohibited_symbols_from_path_with_symbol(data_item[A.CT_FNC.NAME][0: min(len(data_item[A.CT_FNC.NAME]), INVENTORY.NAME_MAX_LENTH)])}",
                    )
                )
            return self.create_barcodes(barcode_data, result_directory)
        else:
            raise IncorrectInputFile()

    def drop_note_cache(self) -> None:
        self.keep.sync(True)

    def save_inventory_report_item(
        self, report_file_path: str, inventory_report_item: dict
    ) -> bool:
        row: int = inventory_report_item[A.CT_FNC.ROW]
        column: int = inventory_report_item[A.CT_FNC.QUANTITY_COLUMN]
        value: str = str(inventory_report_item[A.CT_FNC.QUANTITY])
        if self.is_excel_old(report_file_path):
            src_xls = xlrd.open_workbook(report_file_path)
            dest_xls = copy(src_xls)
            sheet = dest_xls.get_sheet(0)
            sheet.write(row, column, value)
            dest_xls.save(report_file_path)
        else:
            workbook = self.workbook_map[report_file_path]
            sheet = workbook.active
            sheet.cell(row, column).value = value
            workbook.save(filename=report_file_path)
        return True

    def get_inventory_report_columns(
        self, sheet: Any, old_format: bool = False
    ) -> tuple[int, int, int] | None:
        name_colx: int | None = None
        inventory_number_colx: int | None = None
        quantity_colx: int | None = None

        def normalize(value: str) -> str:
            return value.replace("-", "").replace(nl(), "").lower()

        if old_format:
            for row in range(sheet.nrows):
                for column in range(sheet.ncols):
                    cell_value = str(sheet.cell(rowx=row, colx=column).value)
                    if (
                        cell_value is not None
                        and isinstance(cell_value, str)
                        and len(cell_value) > 0
                    ):
                        cell_value = normalize(cell_value)
                        if name_colx is None:
                            if cell_value == INVENTORY.NAME_COLUMN_NAME:
                                name_colx = column
                        elif quantity_colx is None:
                            if cell_value == INVENTORY.QUANTITY_COLUMN_NAME:
                                quantity_colx = column
                        elif inventory_number_colx is None:
                            if cell_value == INVENTORY.NUMBER_COLUMN_NAME:
                                inventory_number_colx = column
                                return (name_colx, inventory_number_colx, quantity_colx)
        else:
            max_row = sheet.max_row
            max_col = sheet.max_column
            for row in sheet.iter_rows(
                min_row=1, min_col=1, max_row=max_row, max_col=max_col
            ):
                for cell in row:
                    cell_value = cell.value
                    if (
                        cell_value is not None
                        and isinstance(cell_value, str)
                        and len(cell_value) > 0
                    ):
                        cell_value = normalize(cell_value)
                        if name_colx is None:
                            if cell_value == INVENTORY.NAME_COLUMN_NAME:
                                name_colx = cell.column
                        elif quantity_colx is None:
                            if cell_value == INVENTORY.QUANTITY_COLUMN_NAME:
                                quantity_colx = cell.column
                        elif inventory_number_colx is None:
                            if cell_value == INVENTORY.NUMBER_COLUMN_NAME:
                                inventory_number_colx = cell.column
                                return (name_colx, inventory_number_colx, quantity_colx)
        return None

    def close_inventory_report(self, report_file_path: str) -> None:
        if report_file_path in self.workbook_map:
            self.workbook_map[report_file_path].close()
            del self.workbook_map[report_file_path]

    def is_excel_old(self, file_path: str) -> bool:
        return A.PTH.get_extension(file_path) == A.CT_F_E.EXCEL_OLD

    def inventory_report_columns_is_exists(self, report_file_path: str) -> bool:
        old_format: bool = self.is_excel_old(report_file_path)
        return (
            self.get_inventory_report_columns(
                (
                    xlrd.open_workbook(report_file_path).sheet_by_index(0)
                    if old_format
                    else openpyxl.load_workbook(
                        report_file_path, data_only=True, read_only=True
                    ).active
                ),
                old_format,
            )
            is not None
        )

    def inventory_report(
        self, report_file_path: str, open_for_edit: bool = False
    ) -> list[dict] | None:
        old_format: bool = self.is_excel_old(report_file_path)
        sheet = None
        if old_format:
            book = xlrd.open_workbook(report_file_path, formatting_info=True)
            sheet = book.sheet_by_index(0)
            name_colx: int | None = None
            inventory_number_colx: int | None = None
            quantity_colx: int | None = None
            inventory_report_columns = self.get_inventory_report_columns(
                sheet, old_format
            )
            if inventory_report_columns is None:
                return None
            else:
                (
                    name_colx,
                    inventory_number_colx,
                    quantity_colx,
                ) = inventory_report_columns
            result_data: list[dict[str, str]] = []
            for rowx in range(sheet.nrows):
                name_cell = sheet.cell(rowx=rowx, colx=name_colx)
                inventory_number_cell = sheet.cell(
                    rowx=rowx, colx=inventory_number_colx
                )
                inventory_number_cell_xf = book.xf_list[inventory_number_cell.xf_index]
                inventory_number_cell_format = book.format_map[
                    inventory_number_cell_xf.format_key
                ]
                format_str = inventory_number_cell_format.format_str
                if inventory_number_cell.value != "" and (
                    format_str == "General" or len(format_str) > 1
                ):
                    if all(x == "0" for x in format_str):
                        format_str_len: int = len(format_str)
                        inventory_number = str(int(inventory_number_cell.value))
                        inventory_number = (
                            "0" * (format_str_len - len(inventory_number))
                            + inventory_number
                        )
                    else:
                        inventory_number = inventory_number_cell.value
                    name: str = name_cell.value
                    name_value: str = name
                    name = name.lower()
                    if len(name) > 0 and name not in [
                        INVENTORY.NAME_COLUMN_NAME,
                        INVENTORY.NUMBER_COLUMN_NAME,
                    ]:
                        quantity = sheet.cell(rowx=rowx, colx=quantity_colx).value
                        result_data.append(
                            {
                                A.CT_FNC.NAME: name_value,
                                A.CT_FNC.INVENTORY_NUMBER: inventory_number,
                                A.CT_FNC.QUANTITY: quantity,
                                A.CT_FNC.ROW: rowx,
                                A.CT_FNC.NAME_COLUMN: name_colx,
                                A.CT_FNC.INVENTORY_NUMBER_COLUMN: inventory_number_colx,
                                A.CT_FNC.QUANTITY_COLUMN: quantity_colx,
                            }
                        )
        else:
            workbook = None
            if open_for_edit:
                if report_file_path not in self.workbook_map:
                    workbook = openpyxl.load_workbook(filename=report_file_path)
                    self.workbook_map[report_file_path] = workbook
                else:
                    workbook = self.workbook_map[report_file_path]
                sheet = workbook.active
            else:
                with open(report_file_path, "rb") as file:
                    in_memory_file = io.BytesIO(file.read())
                workbook = openpyxl.load_workbook(
                    in_memory_file, data_only=True, read_only=True
                )
                sheet = workbook.active
            inventory_report_columns = self.get_inventory_report_columns(
                sheet, old_format
            )
            if inventory_report_columns is None:
                return None
            else:
                (
                    name_colx,
                    inventory_number_colx,
                    quantity_colx,
                ) = inventory_report_columns
            result_data: list[dict[str, str]] = []
            for y, row in enumerate(sheet.rows):
                name: str | None = None
                name_value: str | None = None
                inventory_number: str | None = None
                quantity: str | None = None
                for x, cell in enumerate(row):
                    if x in [
                        name_colx - 1,
                        inventory_number_colx - 1,
                        quantity_colx - 1,
                    ]:
                        if x == name_colx - 1:
                            name_value = cell.value
                            if (
                                name_value is not None
                                and isinstance(name_value, str)
                                and len(name_value) > 5
                            ):
                                if name_value.lower() != INVENTORY.NAME_COLUMN_NAME:
                                    name = name_value
                                else:
                                    break
                            else:
                                break
                        if x == inventory_number_colx - 1:
                            format_str: str = cell.number_format
                            inventory_number = cell.value
                            if all(x == "0" for x in format_str):
                                inventory_number = str(inventory_number)
                                format_str_len: int = len(format_str)
                                inventory_number = (
                                    "0" * (format_str_len - len(inventory_number))
                                    + inventory_number
                                )
                        if x == quantity_colx - 1:
                            quantity = cell.value
                            if (
                                nn(name)
                                and inventory_number is not None
                                and quantity is not None
                            ):
                                result_data.append(
                                    {
                                        A.CT_FNC.NAME: name,
                                        A.CT_FNC.INVENTORY_NUMBER: inventory_number,
                                        A.CT_FNC.QUANTITY: quantity,
                                        A.CT_FNC.ROW: y + 1,
                                        A.CT_FNC.NAME_COLUMN: name_colx,
                                        A.CT_FNC.INVENTORY_NUMBER_COLUMN: inventory_number_colx,
                                        A.CT_FNC.QUANTITY_COLUMN: quantity_colx,
                                    }
                                )
                                break
        return result_data

    def fill_template(
        self,
        date_now_string: str,
        name: str,
        tab_number: str,
        web_site_name: str,
        web_site: str,
        email_address: str,
        pc: LoginPasswordPair | None = None,
        polibase: LoginPasswordPair | None = None,
        email: LoginPasswordPair | None = None,
    ) -> MailMerge:
        paswword_caption_values = ["компьютер", "Polibase", "электронную почту"]
        if pc is None:
            pc = LoginPasswordPair()
        #
        polibase = polibase or pc
        email = email or pc
        #
        internal_email = email.login.find(web_site_name) != -1
        if not internal_email:
            email_address = email.login[email.login.find("@") + 1 :]
        #
        template_file_name: str | None = None
        if polibase.login == pc.login:
            template_file_name = "Template1.docx"
        else:
            template_file_name = "Template2.docx"
        #
        pc_password = pc.password
        polibase_password = polibase.password or pc_password
        email_password = email.password or polibase_password
        #
        password_caption_items = [[], [], []]
        password_text_items = ["", "", ""]
        #
        password_caption_values_len = len(paswword_caption_values)
        password_caption_items[0].append(paswword_caption_values[0])
        password_text_items[0] = pc_password
        if polibase_password == pc_password:
            password_caption_items[0].append(paswword_caption_values[1])
        else:
            password_caption_items[1].append(paswword_caption_values[1])
            password_text_items[1] = polibase_password
        if internal_email:
            if email_password == polibase_password or email_password == pc_password:
                password_caption_items[0].append(paswword_caption_values[2])
            else:
                password_caption_items[2].append(paswword_caption_values[2])
                password_text_items[2] = email_password
        password_caption: str = ""
        password_text: str = ""
        if (
            len(password_caption_items[0])
            == password_caption_values_len - int(not internal_email)
            and pc_password is not None
        ):
            password_caption = nl("Ваш пароль:", reversed=True)
            password_text = nl(pc_password, reversed=True)
        else:
            for i in range(password_caption_values_len):
                password_caption_item = password_caption_items[i]
                password_caption_item_len = len(password_caption_item)
                if password_caption_item_len > 0:
                    password = password_text_items[i]
                    if password is not None and password != "":
                        password_caption += nl(
                            j(
                                (
                                    "Пароль для входа в ",
                                    j(password_caption_item, nl(",")),
                                    ":",
                                )
                            )
                        )
                        password_text += password
                        password_text += nl(count=password_caption_item_len)
            if password_caption == "":
                template_file_name = "Template0.docx"

        template_file_path: str = A.PTH.join(
            A.PTH.FACADE.SERVICE_FILES(SD.standalone_name), template_file_name
        )

        def fill_template_internal(
            template_file_path: str,
            name: str,
            tab_number: str,
            email_address: str,
            web_site: str,
            date_now_string: str,
            password_caption: str,
            password_text: str,
            pc: LoginPasswordPair,
            polibase: LoginPasswordPair,
            email: LoginPasswordPair,
        ) -> MailMerge:
            document = MailMerge(template_file_path)
            document.merge(
                Name=name,
                Date=date_now_string,
                TabNumber=tab_number,
                Login=pc.login,
                WebSite=web_site,
                EMailAddress=email_address,
                EMailLogin=email.login,
                PolibaseLogin=polibase.login,
                PasswordCaption=password_caption,
                PasswordText=password_text,
            )
            return document

        return fill_template_internal(
            template_file_path,
            name,
            tab_number,
            email_address,
            web_site,
            date_now_string,
            password_caption,
            password_text,
            pc,
            polibase,
            email,
        )

    def create(
        self,
        path: str,
        date_now_string: str,
        web_site_name: str,
        web_site: str,
        email_address: str,
        name: str,
        tab_number: str,
        pc: LoginPasswordPair,
        polibase: LoginPasswordPair,
        email: LoginPasswordPair,
    ) -> None:
        self.fill_template(
            date_now_string,
            name,
            tab_number,
            web_site_name,
            web_site,
            email_address,
            pc,
            polibase,
            email,
        ).write(path)
