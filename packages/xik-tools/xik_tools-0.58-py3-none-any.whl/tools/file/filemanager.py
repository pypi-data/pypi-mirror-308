import os
import pickle
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import pandas as pd
import logging

class FileManager:
    def __init__(self):
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)

    @staticmethod
    def make_dir(folder_path):
        relative_path = folder_path.split("/")
        current_path = os.getcwd()
        for folder_name in relative_path:
            current_path = os.path.join(current_path, folder_name)
            if not os.path.exists(current_path):
                os.makedirs(current_path)
                logging.info(f"폴더 '{folder_name}'가 생성되었습니다.")

    @staticmethod
    def delete_dir(folder_path):
        try:
            if os.path.exists(folder_path):
                for root, dirs, files in os.walk(folder_path, topdown=False):
                    for name in files:
                        file_path = os.path.join(root, name)
                        os.remove(file_path)
                        logging.info(f"파일 '{file_path}'가 삭제되었습니다.")
                    for name in dirs:
                        dir_path = os.path.join(root, name)
                        os.rmdir(dir_path)
                        logging.info(f"디렉토리 '{dir_path}'가 삭제되었습니다.")
                os.rmdir(folder_path)
                logging.info(f"The folder at {folder_path} has been deleted.")
            else:
                logging.warning(f"The folder at {folder_path} does not exist.")
        except Exception as e:
            logging.error(f"Error: {e}")

    @staticmethod
    def save_file(data, filename="backup"):
        with open(f"{filename}.pickle", "wb") as file:
            pickle.dump(data, file)
        logging.info(f"{filename}.pickle 파일에 데이터가 저장되었습니다.")
        return data

    @staticmethod
    def load_file(filename="backup"):
        with open(f"{filename}.pickle", 'rb') as file:
            data = pickle.load(file)
        logging.info(f"{filename}.pickle 파일에서 데이터가 로드되었습니다.")
        return data

    @staticmethod
    def get_datetime_info(include_time=True):
        format_str = "%Y-%m-%d_%H-%M-%S" if include_time else "%Y-%m-%d"
        return datetime.now().strftime(format_str)

    @staticmethod
    def get_name_from_url(url):
        if url[-1] == '/':
            url = url[:-1]
        name = url.rsplit('/', 1)[-1]
        logging.info(f"URL '{url}'에서 이름 '{name}'이(가) 추출되었습니다.")
        return name

    @staticmethod
    def dict_to_excel(dict_data, file_name="dictToExcel", sheet_name="sheet1", orient_idx=True):
        orient = 'columns'
        if orient_idx: orient = 'index'

        df = pd.DataFrame.from_dict(dict_data, orient=orient).reset_index().rename(columns={'index': 'model'})

        try:
            wb = load_workbook(filename=f"{file_name}.xlsx")
            ws = wb.create_sheet(title=sheet_name)
            logging.info(f"기존 엑셀 파일 '{file_name}.xlsx'에서 시트 '{sheet_name}'이(가) 생성되었습니다.")
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            logging.info(f"새 엑셀 파일 '{file_name}.xlsx'가 생성되었습니다.")

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        wb.save(f"{file_name}.xlsx")
        logging.info(f"데이터가 {file_name}.xlsx 파일의 {sheet_name} 시트에 저장되었습니다.")

    @staticmethod
    def df_to_excel(df_data, file_name="dfToExcel", sheet_name="sheet1", mode='w'):
        # Create Excel writer
        with pd.ExcelWriter(file_name, engine='openpyxl', mode=mode) as writer:
            # Write each DataFrame to a different sheet
            df_data.to_excel(writer, sheet_name=sheet_name, index=False)
            logging.info(f"DataFrame이 {file_name}.xlsx 파일의 {sheet_name} 시트에 저장되었습니다.")


class DataComparator:
    @staticmethod
    def is_equal_each_columns(ds1: pd.Series, ds2: pd.Series):
        """
        # 예시 데이터
        data1 = {'A': 1, 'B': 2, 'C': 3}
        data2 = {'A': 1, 'B': 5, 'C': 3}

        # Pandas Series 생성
        series1 = pd.Series(data1)
        series2 = pd.Series(data2)

        # DataComparator 인스턴스 생성
        comparator = DataComparator()

        # 메서드 호출
        comparator.is_equal_each_columns(series1, series2)
        """
        # 값이 같은지 여부 확인
        are_equal = ds1.equals(ds2)
        # 값이 다른 항목 확인
        diff_positions = (ds1 != ds2)
        # 차이가 있는 항목 출력
        diff_items = ds1.index[diff_positions].tolist()
        
        logging.info(f"두 행 간의 값이 같은지 여부: {are_equal}")
        logging.info(f"차이가 있는 항목: {diff_items}")
        print(f"두 행 간의 값이 같은지 여부: {are_equal}")
        print(f"차이가 있는 항목: {diff_items}")
