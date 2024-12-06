############################################################################
#  Author : eunseok, Kim
#  E-mail : es.odysseus@gmail.com
###########################################################################

import json
import string
import random
from typing import Union
import pandas as pd
import numpy as np
import openpyxl as xl
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import *
from openpyxl.styles import colors as color
from openpyxl.styles import borders as border
from openpyxl.chart import *
from openpyxl.utils import dataframe as xdf
from openpyxl.utils import *
from openpyxl.utils.exceptions import *
from ast import literal_eval
from .exception import *
from .lib_path import CPath as lpath
from . import lib_file as lfile
from . import lib_logger as myLogger
from .lib_logger import *

logger = myLogger.get_instance(level_of_file=CmyLogger.DEBUG_LEVEL)


class Cexcel(object):
        
    class Color:
        BLACK = color.COLOR_INDEX[0]
        WHITE = color.COLOR_INDEX[1]
        RED = color.COLOR_INDEX[2]
        GREEN = color.COLOR_INDEX[3]
        BLUE = color.COLOR_INDEX[4]
        YELLOW = color.COLOR_INDEX[5]
        DARKRED = color.COLOR_INDEX[8]
        DARKGREEN = color.COLOR_INDEX[9]
        DARKBLUE = color.COLOR_INDEX[12]
        DARKYELLOW = color.COLOR_INDEX[19]

    JSON_POSTFIX='_json_data.json'

    _excel_path_ = None
    _workbook_ = None
    _sheets_ = dict()
    _def_side_ = Side(style=border.BORDER_THIN, color=color.BLACK)
    _def_border_ = Border(left=_def_side_, top=_def_side_, right=_def_side_, bottom=_def_side_)
    _def_align_ = Alignment(horizontal="center", vertical="center")
    _def_style_ = NamedStyle(name="def_style", border=_def_border_, alignment=_def_align_)
    
    ########################################
    # Public Function Definition
    ########################################
    @staticmethod
    def create_simple_df(data: Union[dict, list], columns: list):
        """_summary_
        Create DataFrame data base on Dictionary and List data

        Args:
            data (Union[dict, list]): 
                - dict  {
                            'row1': [ data1, data2, data3, data4, data5 ],
                            'row2': [ data1, data2, data3, data4, data5 ]
                        }
                - list  [ data1, data2, data3, data4, data5 ]
                        ~~~ Or ~~~
                        [
                            [ data1, data2, data3, data4, data5 ],
                            [ data1, data2, data3, data4, data5 ]
                        ]
            columns (list): [ col1, col2, col3, col4, col5 ]
            
        Raises:
            CENullException: Dict-data is None.
            CENullException: Columns is None.
            
        Returns:
            DataFrame: DataFrame formated data.
        """
        try:
            if data is None:
                raise CENullException("Dict-data is None.")
            if columns is None:
                raise CENullException("Columns is None.")
            
            # Extract indexes ans values
            if isinstance(data, dict) is True:
                indexes = list(data.keys())
                values = list(data.values())
            elif isinstance(data, list) is True:
                indexes = None
                values = data
                if np.array(values).ndim == 1:
                    values = [ values ]
            else:
                raise CENotSupportException(f"Not Supported data-type.({type(data)})")
            
            # Make DataFrame
            return Cexcel.create_df(data=values, columns=columns, indexes=indexes)
        
        except BaseException as e:
            logger.error(e)
            raise e
        
    @staticmethod
    def create_df(data: list, columns: list=None, indexes: list=None):
        """_summary_
        Create DataFrame data base on Dictation data.

        Args:
            data (list): [ data1, data2, data3, data4, data5 ]
                        ~~~ Or ~~~
                        [
                            [ data1, data2, data3, data4, data5 ],
                            [ data1, data2, data3, data4, data5 ]
                        ]
            columns (list): [ col1, col2, col3, col4, col5 ]
                        ~~~ Or ~~~
                        [
                            [ ser1, ser1, ser2, ser2, ser2 ],
                            [ col1, col2, col1, col2, col3 ]
                        ]
            indexes (list, optional): Defaults to None.
                        [ idx1, idx2, idx3, idx4, idx5 ]
                        ~~~ Or ~~~
                        [
                            [ ser1, ser1, ser2, ser2, ser2 ],
                            [ idx1, idx2, idx1, idx2, idx3 ]
                        ]

        Raises:
            CENullException: Dict-data is None.
            CENullException: Indexes is None.
            
        Returns:
            DataFrame: DataFrame formated data.
        """
        try:
            if data is None:
                raise CENullException("Dict-data is None.")
            if isinstance(data, list) is False:
                raise CENotSupportException(f"Not Supported data-type.({type(data)})")
            
            # Extract indexes and values
            values = data
            if np.array(values).ndim == 1:
                values = [ values ]

            if columns is not None:
                if np.array(columns).ndim == 1:
                    columns = [ columns ]
                    
                # Validation check
                if len(columns[0]) != len(values[0]):
                    raise CEInvalidArguments("Missmatched between column's size and data's size.")
                
                # Make Columns
                if len(columns) == 1:
                    columns = columns[0]
                
            if indexes is not None:
                if np.array(indexes).ndim == 1:
                    indexes = [ indexes ]
                
                # Validation check
                if len(indexes[0]) != len(values):
                    raise CEInvalidArguments("Missmatched between indexes's size and data's size.")
                
                # Make Indexes
                if len(indexes) == 1:
                    indexes = indexes[0]
                
            # Make DataFrame
            return pd.DataFrame(values, columns=columns, index=indexes)
            
        except BaseException as e:
            logger.error(e)
            raise e
    
    def create_cell_style(self, font_size: int, font_color: Color,
                          font_bold: bool=False,
                          font_italic: bool=False,
                          _border: Border=_def_border_,
                          _alignment: Alignment=_def_align_,
                          _fill: PatternFill=None):
        """_summary_
        Create Style object for cell.

        Args:
            font_size (int): font size
            font_color (Color): font color
            font_bold (bool, optional): font bold. Defaults to False.
            font_italic (bool, optional): font italic. Defaults to False.
            _border (Border, optional): border style. Defaults to _def_border_.
            _alignment (Alignment, optional): text alignment style. Defaults to _def_align_.
            _fill (PatternFill, optional): cell filling color-style. Defaults to None.

        Raises:
            CENotInitialized: WorkBook is not Initialized.

        Returns:
            NamedStyle: Style object for cell.
        """
        try:
            if self._workbook_ is None:
                raise CENotInitialized("WorkBook is not Initialized.")
            
            _name = self._id_generator_()
            style = NamedStyle(name=_name, font=Font(bold=font_bold, size=font_size, color=font_color, italic=font_italic),
                               border=_border, alignment=_alignment, fill=_fill)
            
            if style.name not in self._workbook_.named_styles:
                self._workbook_.add_named_style(style)
            return style
        
        except BaseException as e:
            logger.error(e)
            raise e
    
    def get_sheet(self, name: str, not_exist_then_create: bool = True):
        sheet = None
        path_excel = None
        
        try:
            if self._workbook_ is None:
                raise CENotInitialized("WorkBook is not initialized.")
            if name is None:
                raise CENullException("Sheet name is None.")
            
            # If already it's exist, then direct return it.
            sheet = self._sheets_.get(name, None)
            if sheet is not None:
                return sheet
            
            # Find worksheet in workbook.
            sheet = self._workbook_[name]
            path_excel = self._excel_path_  # It's need only for already exist sheet.
        except KeyError as e:
            if not_exist_then_create is True:
                # If not exist in workbook, then create sheet.
                sheet = self._workbook_.create_sheet(name)
        except BaseException as e:
            logger.error(e)
            raise e

        if sheet is not None:
            sheet = _Csheet_(sheet, excel_path=path_excel)
            self._sheets_.update( {sheet.name : sheet} )
        return sheet
    
    def get_sheet_names(self):
        try:
            if self._workbook_ is None:
                raise CENotInitialized("WorkBook is not initialized.")
            
            return self._workbook_.sheetnames
        except BaseException as e:
            logger.error(e)
            raise e
        
    def remove_sheet(self, name: str):
        try:
            if self._workbook_ is None:
                raise CENotInitialized("WorkBook is not initialized.")
            if name is None:
                raise CENullException("Sheet name is None.")
            
            Cexcel._remove_sheet_(self._workbook_, name)
            sheet = self._sheets_.get(name, None)
            if sheet is not None:
                del sheet
                if name in list(self._sheets_.keys()):
                    raise CELogicErrorException(f"Can't delete sheet by name.({name})")
        except BaseException as e:
            logger.error(e)
            raise e
        
    def save(self, file_path: str):
        try:
            if self._workbook_ is None:
                raise CENotInitialized("WorkBook is not initialized.")
            if file_path is None:
                raise CEInvalidArguments("file_path is None.")
            
            # Save Excel file.
            self._workbook_.save(file_path)
            
            # Save Json-file for DataFrame talbes.
            json_data = dict()
            path, _ = lpath.split_filename_ext(file_path)
            for name in self.get_sheet_names():
                sheet = self._sheets_[name]
                json_data.update( {name: sheet.json_data} )
                
            lfile.create_json_file(file_path=path + self.JSON_POSTFIX, data=json_data)
        except BaseException as e:
            logger.error(e)
            raise e
        
    def close(self):
        if self._workbook_ is not None:
            for sheet in self._sheets_.values():
                del sheet
            self._sheets_.clear()
            self._workbook_.close()
            self._workbook_ = None
    
    
    ########################################
    # Private Function Definition
    ########################################
    def __init__(self, file_path: str=None, wb: xl.Workbook=None):
        try:
            if file_path is not None:
                wb = xl.load_workbook(file_path)
            elif wb is None:
                wb = xl.Workbook()
                for name in wb.sheetnames:
                    self._remove_sheet_(wb, name)
            
            self._workbook_ = wb
            if self._def_style_.name not in self._workbook_.named_styles:
                self._workbook_.add_named_style(self._def_style_)
            self._sheets_.clear()
            self._excel_path_ = file_path
        except BaseException as e:
            logger.error(e)
            raise e
    
    @staticmethod
    def _id_generator_(size=10):
        try:
            ascii_range = string.ascii_uppercase + string.digits
            return ''.join(random.choice(ascii_range) for _ in range(size))
        except BaseException as e:
            logger.error(e)
            raise e
        
    @staticmethod
    def _remove_sheet_(wb: xl.Workbook, name: str):
        try:
            sheet = None
            if wb is None or name is None:
                raise CENullException("Workbook or Sheet-name is None.")
            
            try:
                sheet = wb[name]
            except KeyError as e:
                sheet = None
            except BaseException as e:
                logger.error(e)
                raise e
                
            if sheet is not None:
                wb.remove( sheet )
        except BaseException as e:
            logger.error(e)
            raise e
        
        
    
            

class _Csheet_(object):
    class MERGE_DIR:
        COLUMNS = 1
        ROWS = 2
    
    _sheet_ = None
    _json_sheet_ = None         # It maintain appended DataFrame-data as Json-data
    _def_side_ = Cexcel._def_side_
    _def_border_ = Cexcel._def_border_
    _def_align_ = Cexcel._def_align_
    _def_style_ = Cexcel._def_style_
    
    ########################################
    # Public Function Definition
    ########################################
    def append_message(self, msg: str, 
                             start_col_name: str, 
                             end_col_name: str, 
                             gap_blank: int=1,
                             gap_post: int=0,
                             border_rows: int=1,
                             style: NamedStyle=_def_style_):
        try:
            if msg is None:
                raise CENullException("Message is None.")
            if start_col_name is None:
                raise CENullException("start_col_name is None.")
            if end_col_name is None:
                raise CENullException("end_col_name is None.")
            
            self._make_title_(start_col_name, end_col_name, msg, gap_blank, border_rows, style)
            for _ in range(0, gap_post, 1):
                self._sheet_['A' + str(self._next_srow_(self._sheet_.max_row))] = ""
                
        except BaseException as e:
            logger.error(e)
            raise e
    
    def append_df(self, df: pd.DataFrame, 
                        title: str=None,
                        start_col_name: str='A',
                        gap_blank: int=1,
                        gap_title: int=1,
                        title_border_rows: int=3,
                        title_style: NamedStyle=_def_style_,
                        skip_column: bool=False,
                        column_style: NamedStyle=None,
                        column_space: list=None,
                        index_style: NamedStyle=_def_style_,
                        base_style: NamedStyle=_def_style_,
                        desp_text: str=None):
        """_summary_
        Append DataFrame formated data to WorkSheet of Excel.

        Args:
            title (str): name of title to be added to worksheet.
            df (pd.DataFrame): source-data to be added to worksheet.
            start_col_name (str, optional): Start column-cell's name. Defaults to 'A'.
            gap_blank (int, optional): Empty-rows count before Appending data. Defaults to 1.
            title_border_rows (int, optional): Title-border's rows count. Defaults to 3.
            title_style (NamedStyle, optional): Cell-Style for title-border. Defaults to _def_style_.
            skip_column (bool, optional): True  => Visible columns-names in excel-file
                                          False => Invisible columns-names in excel-file. Defaults to False.
            column_style (NamedStyle, optional): Cell-Style for columns. Defaults to None.
            index_style (NamedStyle, optional): Cell-Style for indexes. Defaults to _def_style_.
            base_style (NamedStyle, optional): Base Cell-Style for table. Defaults to _def_style_.
            desp_text (str, optional): description message for the table. Defaults to None.
            
        Raises:
            CENullException: DataFrame is None.
            CENullException: DataFrame is None.
            CENullException: WorkSheet or JsonSheet is None.
        """
        def cls_apply_style( srow_num: int, erow_num: int, end_col: str ):
            try:
                scell = chr(ord(start_col_name)+df.index.nlevels) + str(srow_num)
                ecell = end_col + str(erow_num)
                self._apply_style_(scell, ecell, base_style)   # set table-overall
                
                if skip_column is False:
                    ecell = end_col + str(srow_num + df.columns.nlevels - 1)
                    self._apply_style_(scell, ecell, column_style) # set columns
                    srow_num += df.columns.nlevels
                
                scell = chr(ord(start_col_name)) + str(srow_num)
                ecell = chr(ord(start_col_name) + df.index.nlevels - 1) + str(erow_num)
                self._apply_style_(scell, ecell, index_style)  # set indexes
            except BaseException as e:
                logger.error(e)
                raise e
           
        try:
            if df is None:
                raise CENullException("DataFrame is None.")
            if self._sheet_ is None or self._json_sheet_ is None:
                raise CENullException("WorkSheet or JsonSheet is None.")
                
            if column_space is None:
                column_space = [ 1 for _ in range(0, len(df.columns), 1) ]
            end_col_name = chr(ord(start_col_name)+df.index.nlevels+sum(column_space)-1)

            ### Make title
            if title is not None:
                self._make_title_(start_col_name, end_col_name, title, gap_blank, title_border_rows, title_style)
                if skip_column is True:
                    gap_title = 0
                for _ in range(0, gap_title, 1):
                    self._sheet_['A' + str(self._next_srow_(self._sheet_.max_row))] = ""
            
            ### append-rows and register column-part-tables.
            srow_num = self._next_srow_(self._sheet_.max_row)
            pos_cols = self._append_rows_(df, start_col=start_col_name, start_row=srow_num, 
                                          column_space=column_space, skip_column=skip_column)
            erow_num = self._sheet_.max_row
            if skip_column is False:
                self._register_col_tables_( df, srow_num, erow_num, pos_cols )
            
            ### Set Styles for Table-Overall / Columns / Indexes
            cls_apply_style( srow_num, erow_num, end_col_name )
            if desp_text is not None:       # Append description text
                self._sheet_[start_col_name + str(self._next_srow_(self._sheet_.max_row))] = desp_text
                
            ### Append DataFrame as Json-Text
            self._json_sheet_.append_table(title, df, desp_text)
        except BaseException as e:
            logger.error(e)
            raise e
        
    def extract_dfs(self):
        try:
            if self._json_sheet_ is None:
                raise CENullException("JsonSheet is None.")
            
            dfs = list()
            for idx in range(0, self._json_sheet_.table_cnt, 1):
                dfs.append( self._json_sheet_.get_table(idx) )
            return dfs
        except BaseException as e:
            logger.error(e)
            raise e
    
    @property
    def name(self):
        if self._sheet_ is None:
            raise CENullException("WorkSheet is None.")
        return self._sheet_.title
    
    @property
    def json_data(self):
        if self._json_sheet_ is None:
            raise CENullException("JsonSheet is None.")
        return self._json_sheet_.data
    
    ########################################
    # Private Function Definition
    ########################################
    def __init__(self, sheet: Worksheet, excel_path: str=None):
        try:
            if sheet is None:
                raise CENullException("Sheet is None.")
            self._sheet_ = sheet
            self._load_json_data_(excel_path)

        except BaseException as e:
            logger.error(e)
            raise e
    
    def _load_json_data_(self, excel_path: str):
        try:
            json_path = None
            if excel_path is not None:
                path, _ = lpath.split_filename_ext(excel_path)
                json_path = path + Cexcel.JSON_POSTFIX
                
            self._json_sheet_ = _CjsonSheet_(self.name, json_path)
                
        except BaseException as e:
            logger.error(e)
            raise e
    
    def _append_rows_(self, df: pd.DataFrame, start_col: str, start_row: int, column_space: list, 
                      skip_column: bool=False) -> list:
        pos_cols = None
        try:
            if len(column_space) != len(df.columns):
                raise CEInvalidArguments("Missmatched between column_space length & DataFrame.columns length.")
            
            def cls_make_pos_colname():
                pos_list = list()
                for n in range(0, len(df.columns)+df.index.nlevels, 1):
                    if n > df.index.nlevels:
                        # n-th col-pos = sum of column-space elements. (from 0 to n-1)
                        n = sum(column_space[0: n - df.index.nlevels]) + df.index.nlevels
                    pos_list.append( chr(ord(start_col)+n) )
                return pos_list
            
            def cls_merge_colspace(row: int, col_pos: list):
                for i, cells in enumerate(column_space):
                    if cells <= 1:
                        continue
                    scell = col_pos[i + df.index.nlevels] + str(row)
                    ecell = chr(ord(col_pos[i + df.index.nlevels]) + cells -1 ) + str(row)
                    self._sheet_.merge_cells(f"{scell}:{ecell}")
            
            
            pos_cols = cls_make_pos_colname()
            
            ### Append rows to worksheet.
            for idx, r in enumerate(xdf.dataframe_to_rows(df, index=True, header=True)):
                if len(r) == df.index.nlevels and r[0] is None:
                    continue    # if r is empty-seperator-row, then ignore it.
                if skip_column is True and idx < df.columns.nlevels:
                    continue
                
                r = { pos_cols[i]: data[0] if isinstance(data, tuple) is True else data for i,data in enumerate(r) }
                self._sheet_.append(r)
                
                # Merge column-space cells
                row_num = self._sheet_.max_row
                if idx >= df.columns.nlevels -1:    # except upper columns of multi-columns
                    cls_merge_colspace(row_num, pos_cols)
                
                # Merge multilevel-column
                if skip_column is False:
                    if df.columns.nlevels > 1 and idx < df.columns.nlevels - 1:
                        self._auto_merge_cells_(self.MERGE_DIR.COLUMNS, df=df, level=idx, 
                                                start_col_name=pos_cols[df.index.nlevels:], start_row_num=row_num,
                                                skip_column=skip_column, column_space=column_space)
            # Merge multilevel-index
            if df.index.nlevels > 1:
                for level in range(0, df.index.nlevels, 1):
                    self._auto_merge_cells_(self.MERGE_DIR.ROWS, df=df, level=level, 
                                            start_col_name=pos_cols[level], start_row_num=start_row,
                                            skip_column=skip_column)
        except BaseException as e:
            logger.error(e)
            raise e
        return pos_cols
            
    def _auto_merge_cells_(self, direc: MERGE_DIR, df: pd.DataFrame, 
                                 level: int, start_col_name: Union[str, list], start_row_num: int,
                                 skip_column: bool=False, column_space=None):
        class SYM:
            START=1
            END=2
            
        key = None
        start = None
        end = None
        values = None
        now_row = start_row_num

        # Definition of Lambda / Closure function.
        lambda_calc_colspace = lambda i, sym: chr(ord(start_col_name[i])+column_space[i]-1) if sym == SYM.END else start_col_name[i]
        lambda_pick_excel_cell = lambda i, sym: lambda_calc_colspace(i, sym) + str(now_row) if direc == self.MERGE_DIR.COLUMNS else start_col_name + str(now_row+i)
        
        def reset_start(data: tuple, index: int):
            nonlocal key
            nonlocal end
            key = data[level]
            end = None
            return lambda_pick_excel_cell(index, SYM.START)
        
        def merge_cells(start: str, end: str):
            nonlocal key
            if start != end:
                self._sheet_[start] = key
                self._sheet_.merge_cells(f"{start}:{end}")
    
        try:
            # Initialization.
            if direc == self.MERGE_DIR.COLUMNS:
                if isinstance(start_col_name, list) is False:
                    raise CEInvalidArguments(f"When direc is {self.MERGE_DIR.COLUMNS}, start_col_name has to be List type.")
                if column_space is None:
                    raise CENullException(f"When direc is {self.MERGE_DIR.COLUMNS}, need column_space. (It's None)")
                values = df.columns.values
            elif direc == self.MERGE_DIR.ROWS:
                if isinstance(start_col_name, str) is False:
                    raise CEInvalidArguments(f"When direc is {self.MERGE_DIR.ROWS}, start_col_name has to be String type.")
                values = df.index.values
                if skip_column is False:
                    now_row += df.columns.nlevels
            else:
                raise CENotSupportException(f"Not Supported Direction-type.({direc})")
        
            # Start Auto-Merge processing
            for i, tup in enumerate(values):
                if key is None:
                    start = reset_start(tup, i)
                    
                if key != tup[level]:
                    end = lambda_pick_excel_cell(i-1, SYM.END)
                    merge_cells(start, end)
                    start = reset_start(tup, i)
                    
                if i == len(values)-1 and key is not None:
                    end = lambda_pick_excel_cell(i, SYM.END)
                    merge_cells(start, end)
                    key = None
        
        except BaseException as e:
            logger.error(e)
            raise e
    
    def _register_col_tables_(self, df: pd.DataFrame, srow_num: int, erow_num: int, pos_cols: list ):
        try:
            if pos_cols is None:
                raise CEInvalidArguments("Need pos_cols list But, it's None.")
            
            start_row = srow_num + df.columns.nlevels - 1
            end_row = erow_num
            col_range = pos_cols[df.index.nlevels: ]
            
            for col_name in col_range:
                scell = col_name + str(start_row)
                ecell = col_name + str(end_row)
                parts = Table(displayName=Cexcel._id_generator_(), ref=f"{scell}:{ecell}")
                self._sheet_.add_table( parts )
        except BaseException as e:
            logger.error(e)
            raise e
    
    def _apply_style_(self, start_cell: str, end_cell: str, style: NamedStyle):
        try:
            if style is None:
                logger.info("Style is None.")
                return
            
            if self._sheet_ is None:
                raise CENullException("WorkSheet is None.")
            if start_cell is None or end_cell is None:
                raise CENullException("Start/End cell name is None.")

            min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
            for cells in self._sheet_.iter_rows(min_row, max_row, min_col, max_col):
                for cell in cells:
                    cell.style = style.name
        except BaseException as e:
            logger.error(e)
            raise e
            
    def _make_title_(self, start_col: str, end_col: str, message: str, blank_rows: int, border_gap: int, style: NamedStyle):
        try:
            if style is None:
                raise CEInvalidArguments("Title-style is None.")
            
            border_gap = border_gap - 1
            srow_num = self._next_srow_(self._sheet_.max_row) + blank_rows
            scell = start_col + str(srow_num)
            ecell = end_col + str(srow_num + border_gap)
            self._sheet_[ecell] = self._sheet_[scell] = message
            
            self._sheet_[scell].style = style.name
            self._sheet_.merge_cells(f"{scell}:{ecell}")
        except BaseException as e:
            logger.error(e)
            raise e
    
    @staticmethod
    def _next_srow_(gap_size: int):
        return gap_size if gap_size == 1 else gap_size+1
    

class _CjsonSheet_(object):
    
    #############################
    # Reference
    #   - https://towardsdatascience.com/all-pandas-json-normalize-you-should-know-for-flattening-json-13eae1dfb7dd
    #############################
    class KEY:
        SHEET='sheet_name'
        TITLE='title'
        DESP='desp'
        DATA='data'
        TABLES='tables'
        
    FORMAT_TYPE = 'index'       # valid values: 'split', 'records', 'index', 'columns', 'values', 'table'
    
    _data_ = dict()
    _table_cnt_ = 0
    
    @property
    def data(self):
        return self._data_
    
    @property
    def table_cnt(self):
        return self._table_cnt_

    
    def append_table(self, title: str, df: pd.DataFrame, desp: str=None):
        try:
            if df is None:
                raise CEInvalidArguments('DataFrame is None.')
            if title is None:
                title = "No title"
            
            json_data = {self.KEY.TITLE: title,
                         self.KEY.DESP: desp,
                         self.KEY.DATA: json.loads(df.to_json(orient=self.FORMAT_TYPE))}
            self.data[self.KEY.TABLES].update( {str(self._table_cnt_): json_data} )
            self._table_cnt_ += 1
            
        except BaseException as e:
            logger.error(e)
            raise e

    def get_table(self, idx: int):
        try:
            if idx is None:
                raise CEInvalidArguments("index num is None.")
            if idx >= self.table_cnt:
                raise CEInvalidArguments(f"index num({idx}) is over than table-can({self.table_cnt}).")
            
            table_info = self.data[self.KEY.TABLES][str(idx)]
            title = table_info[self.KEY.TITLE]
            desp = table_info[self.KEY.DESP]
            data = table_info[self.KEY.DATA]
            df = self._make_df_(data)

            return (title, df, desp)
        except BaseException as e:
            logger.error(e)
            raise e

    def __init__(self, name: str, json_path: str=None):
        try:
            if name is None:
                raise CEInvalidArguments('Sheet name is None.')
            
            self.data.clear()
            if json_path is None:
                self.data.update( {self.KEY.SHEET: name} )
                self.data.update( {self.KEY.TABLES: dict()} )
                return
            
            data = lfile.read_json_file(json_path, with_annotation=True)
            if data is None:
                raise CENotFoundException(f"Can not find in {json_path}")

            self._data_, self._table_cnt_ = self._check_validation_( name, data.get(name, None) )
            
        except BaseException as e:
            logger.error(e)
            raise e

    def _make_df_(self, data: dict):
        def cls_make_list_idx_col(target: dict):
            if self._is_multiple_idx_col_(target) is True:
                key_tuples = [ literal_eval(key) for key in target.keys() ]
                return list(map(list, zip(*key_tuples)))
            else:
                return list(target.keys())
            
        try:
            if self.FORMAT_TYPE != 'index':     # Can not supported multiple indexes & columns
                return pd.read_json(data, orient=self.FORMAT_TYPE)
            
            #### It's Support multiple indexes & columns
            # make indexes  
            indexes = cls_make_list_idx_col(data)
            
            columns = None
            values = list()
            for record in data.values():
                # make values
                values.append( list(record.values()) )
                if columns is not None:
                    continue
                
                # make columns
                columns = cls_make_list_idx_col(record)
            
            return Cexcel.create_df(data=values, columns=columns, indexes=indexes)

        except BaseException as e:
            logger.error(e)
            raise e
    
    def _is_multiple_idx_col_(self, data: dict):
        try:
            # check it's tuple or not.
            return isinstance(literal_eval( list(data.keys())[0] ), tuple)
        except SyntaxError as e:
            logger.info(e)
            return False
        except ValueError as e:
            logger.info(e)
            return False
        except BaseException as e:
            logger.error(e)
            raise e
        
    def _check_validation_(self, sheet_name: str, data: dict):
        try:
            if self.KEY.SHEET not in data:
                raise CEInvalidArguments(f"Can not find KEY.({self.KEY.SHEET})")
            if self.KEY.TABLES not in data:
                raise CEInvalidArguments(f"Can not find KEY.({self.KEY.TABLES})")
            if data[self.KEY.SHEET] != sheet_name:
                raise CEMissMatchedException(f"Missmatch sheet-name.(name=> {sheet_name}, {data[self.KEY.SHEET]})")
            
            table_cnt = len(data[self.KEY.TABLES])
            return data, table_cnt
        except BaseException as e:
            logger.error(e)
            raise e
